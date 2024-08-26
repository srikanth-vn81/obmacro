import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# Function to transform the "VPO No" column (for Excel processing)
def transform_vpo_no(vpo_no):
    if isinstance(vpo_no, str):
        if vpo_no.startswith('8'):
            return vpo_no[:8]
        elif vpo_no.startswith('D'):
            return 'P' + vpo_no[1:-3]
    return vpo_no

# Function to convert 'PCD' column from float to datetime (for Excel processing)
def convert_to_date(x):
    try:
        if x and x != '':
            x = str(int(float(x)))  # Remove decimal and convert to string
            return pd.to_datetime(x, format='%Y%m%d', errors='coerce')
        return pd.NaT
    except:
        return pd.NaT

# Function to process the uploaded Excel file
def process_excel(file):
    excel_data = pd.ExcelFile(file)
    data = pd.read_excel(excel_data, sheet_name='Sheet1')

    # List of columns to drop
    columns_to_drop = ['CBU', 'Buyer','Buyer Division Code', 'Cust Style No', 'Product Group',
                       'Style Category', 'Garment Fabrication', 'Lead Factory', 'Prod Warehouse',
                       'Max CO Sts', 'Delivery Term', 'Color Code', 'FOB Date', 'Max Departure Date - CO',
                       'Cum Wash Rev Qty', 'Cum Wash Rev Rej Qty', 'Remaining Qty', 'Allocated Qty',
                       'Invoiced Qty', 'FOB Price', 'FOB after discount', 'SMV', 'Planned SAH',
                       'Costing Efficiency', 'CO Responsible', 'CO Create Min Date', 'CO Create Max Date',
                       'Drop Dead Date', 'AOQ', 'Type', 'Projection Ref']

    # Drop the specified columns
    data_cleaned = data.drop(columns=columns_to_drop)

    # Filter data
    data_cleaned = data_cleaned[data_cleaned['Group Tech Class']=="BELUNIQLO"]

    # Apply the transformation function to the "VPO No" column
    data_cleaned['PO'] = data_cleaned['VPO No'].apply(transform_vpo_no)

    # Ensure 'Production Plan ID' column exists
    if 'Production Plan ID' not in data_cleaned.columns:
        data_cleaned['Production Plan ID'] = None

    # Update 'Production Plan ID' column for rows that are blank
    data_cleaned['Production Plan ID'] = data_cleaned.apply(
        lambda row: (
            str(row['PO']) if pd.isna(row['Production Plan ID']) and str(row['PO']).startswith('8') else
            "Season-23" if (pd.isna(row['Production Plan ID']) or row['Production Plan ID'] == '') and str(row['Season'])[-2:] == '23' else
            row['Production Plan ID']
        ),
        axis=1
    )

    # Convert specific columns to text
    columns_to_convert = ['Min CO Sts', 'Order placement date', 'PCD']
    for column in columns_to_convert:
        if column in data_cleaned.columns:
            data_cleaned[column] = data_cleaned[column].astype(str)

    # Convert 'PCD' column to datetime
    data_cleaned['PCD'] = data_cleaned['PCD'].apply(convert_to_date)

    return data_cleaned

# Function to convert date columns from object to datetime format (for CSV processing)
def convert_dates_to_datetime(df, date_columns):
    for col in date_columns:
        df[col] = pd.to_datetime(df[col], format='%m/%d/%Y', errors='coerce')
    return df

# Function to process the uploaded CSV file
def process_csv(file):
    new_csv_data = pd.read_csv(file)

    # List of columns to keep
    columns_to_keep = ['Production Plan ID', 'Main Sample Code', 'Season', 'Year', 
                       'Production Plan Type Name', 'EXF', 'Contracted Date', 
                       'Requested Wh Date', 'Business Unit', 'PO Order NO']

    # Check if these columns are present in the new data
    available_columns = [col for col in columns_to_keep if col in new_csv_data.columns]

    # Drop all other columns except the ones to keep
    new_csv_data_cleaned = new_csv_data.loc[:, available_columns].copy()

    # Rename the column 'PO Order NO' to 'PO'
    new_csv_data_cleaned.rename(columns={'PO Order NO': 'PO'}, inplace=True)

    # List of date columns to convert
    date_columns = ['EXF', 'Contracted Date', 'Requested Wh Date']

    # Convert the date columns
    new_csv_data_cleaned = convert_dates_to_datetime(new_csv_data_cleaned, date_columns)

    return new_csv_data_cleaned

# Function to process the third Excel file (RFID Gihan)
def process_rfid_excel(file):
    data_2_original = pd.read_excel(file, sheet_name='sheet1')

    # Unmerge the 'DO No./Product No.' column by forward filling the values
    data_2_original['DO No./Product No.'] = data_2_original['DO No./Product No.'].ffill()

    # Apply the condition to 'Set Detail'
    data_2_original['Set Detail'] = data_2_original.apply(
        lambda row: row['Set Detail'] if pd.notnull(row['Set Detail']) and row['Set Detail'] != '-' else row['Set Code'],
        axis=1
    )

    # Select the 'PCS' values for Order Quantity, Packing Quantity, Loading Quantity, and Inventory Quantity
    data_2_original['Order Quantity'] = data_2_original.iloc[:, 4]  # PCS column for Order Quantity
    data_2_original['Packing Quantity'] = data_2_original.iloc[:, 6]  # PCS column for Packing Quantity
    data_2_original['Loading Quantity'] = data_2_original.iloc[:, 8]  # PCS column for Loading Quantity
    data_2_original['Inventory Quantity'] = data_2_original.iloc[:, 10]  # PCS column for Inventory Quantity

    # Drop the first row which contains the repeated headers
    data_2_final = data_2_original.drop(index=0)

    # Reset the index of the cleaned dataframe
    data_2_final.reset_index(drop=True, inplace=True)

    # Remove unnamed columns
    columns_to_keep = ['DO No./Product No.', 'Set Code', 'Set Detail', 'Order Quantity', 'Packing Quantity', 'Loading Quantity', 'Inventory Quantity']
    data_2_cleaned_final = data_2_final.loc[:, columns_to_keep]

    # Convert specified fields to integers
    data_2_cleaned_final.loc[:, 'Order Quantity'] = data_2_cleaned_final['Order Quantity'].astype(int)
    data_2_cleaned_final.loc[:, 'Packing Quantity'] = data_2_cleaned_final['Packing Quantity'].astype(int)
    data_2_cleaned_final.loc[:, 'Loading Quantity'] = data_2_cleaned_final['Loading Quantity'].astype(int)
    data_2_cleaned_final.loc[:, 'Inventory Quantity'] = data_2_cleaned_final['Inventory Quantity'].astype(int)

    # Create a new field 'Color Code' by taking the first 2 digits from the 'Set Code' field and convert it to text
    data_2_cleaned_final.loc[:, 'Color Code'] = data_2_cleaned_final['Set Code'].str[:2].astype(str)

    # Create a new field 'Pack Method' based on the first two characters of 'Set Code', handle NaN values
    data_2_cleaned_final.loc[:, 'Pack Method'] = data_2_cleaned_final['Set Code'].fillna('').apply(
        lambda x: 'AST' if x[:2].isalpha() else ('1SL' if x[:2].isdigit() else ' ')
    )

    return data_2_cleaned_final

# Function to update 'Production Plan ID' in the OB_clean dataframe based on the SPL_clean dataframe
def update_production_plan_id(ob_clean_df, spl_clean_df):
    ob_clean_df['Production Plan ID'] = ob_clean_df.apply(
        lambda row: row['Production Plan ID'] if pd.notnull(row['Production Plan ID']) and row['Production Plan ID'] != '' else 
        spl_clean_df.loc[spl_clean_df['PO'] == row['PO'], 'Production Plan ID'].values[0] 
        if len(spl_clean_df.loc[spl_clean_df['PO'] == row['PO'], 'Production Plan ID'].values) > 0 else 'N/A', 
        axis=1
    )
    return ob_clean_df

# Function to merge the two dataframes on 'Production Plan ID'
def merge_dataframes(ob_clean_final_df, spl_clean_df):
    ob_clean_final_df['Production Plan ID'] = ob_clean_final_df['Production Plan ID'].astype(str)
    spl_clean_df['Production Plan ID'] = spl_clean_df['Production Plan ID'].astype(str)
    merged_df_corrected = pd.merge(ob_clean_final_df, spl_clean_df, on='Production Plan ID', how='left')
    return merged_df_corrected

# Function to perform final calculations and add columns
def perform_final_calculations(merged_df_corrected):
    # Check if 'POWH-PLN' exists and convert to datetime if it does
    if 'POWH-PLN' in merged_df_corrected.columns:
        merged_df_corrected['POWH-PLN'] = pd.to_datetime(merged_df_corrected['POWH-PLN'], errors='coerce')
    else:
        # Add 'POWH-PLN' as an empty column
        merged_df_corrected['POWH-PLN'] = pd.NaT

    # Ensure 'Requested Wh Date' is converted to datetime if it exists
    if 'Requested Wh Date' in merged_df_corrected.columns:
        merged_df_corrected['Requested Wh Date'] = pd.to_datetime(merged_df_corrected['Requested Wh Date'], errors='coerce')

    # Adding other empty columns
    merged_df_corrected['MODE'] = ''
    merged_df_corrected['EXF-PLN'] = ''
    merged_df_corrected['ETD-PLN'] = ''
    merged_df_corrected['Factory – Remarks'] = ''
    merged_df_corrected['Delays'] = np.nan
    merged_df_corrected['Delay/Early'] = ''

    # Calculate percentages
    merged_df_corrected['Cut%'] = ((merged_df_corrected['Cum Cut Qty'] / merged_df_corrected['CO Qty']) * 100).round(2)
    merged_df_corrected['Sewin%'] = ((merged_df_corrected['Cum Sew In Qty'] / merged_df_corrected['CO Qty']) * 100).round(2)
    merged_df_corrected['Sewin Rej%'] = ((merged_df_corrected['Cum Sew In Rej Qty'] / merged_df_corrected['Cum Sew In Qty']) * 100).round(2)
    merged_df_corrected['Sewout%'] = ((merged_df_corrected['Cum SewOut Qty'] / merged_df_corrected['CO Qty']) * 100).round(2)
    merged_df_corrected['Sewout Rej%'] = ((merged_df_corrected['Cum Sew Out Rej Qty'] / merged_df_corrected['Cum SewOut Qty']) * 100).round(2)
    merged_df_corrected['CTN%'] = ((merged_df_corrected['Cum CTN Qty'] / merged_df_corrected['CO Qty']) * 100).round(2)
    merged_df_corrected['Del%'] = ((merged_df_corrected['Delivered Qty'] / merged_df_corrected['CO Qty']) * 100).round(2)

    # Calculate 'Delays'
    if 'Requested Wh Date' in merged_df_corrected.columns and 'POWH-PLN' in merged_df_corrected.columns:
        valid_dates = merged_df_corrected['POWH-PLN'].notna() & merged_df_corrected['Requested Wh Date'].notna()
        if valid_dates.any():
            merged_df_corrected.loc[valid_dates, 'Delays'] = (merged_df_corrected.loc[valid_dates, 'Requested Wh Date'] - merged_df_corrected.loc[valid_dates, 'POWH-PLN']).dt.days

    # Create 'Delay/Early' column based on the condition using numpy where
    if 'Delays' in merged_df_corrected.columns:
        merged_df_corrected['Delay/Early'] = np.where(merged_df_corrected['Delays'] > 0, "Delay", "No Delay")

    # Apply the final filter
    merged_df_corrected = merged_df_corrected[(merged_df_corrected['Production Plan ID'] != '0') & (merged_df_corrected['Production Plan ID'] != 'Season-23')]

    return merged_df_corrected

# Function to add 'Color Code' to merged data based on 'Color Name'
def add_color_code(merged_df_corrected):
    # Create a new field 'Color Code' by taking the first 2 digits from the 'Color Name' field and convert it to text
    if 'Color Name' in merged_df_corrected.columns:
        merged_df_corrected['Color Code'] = merged_df_corrected['Color Name'].astype(str).str[:2]
    return merged_df_corrected

# Function to perform final merge with RFID data and add the 'Status' column
def final_merge_and_status(merged_data, rfid_data):
    # Convert 'Color Code' to string type in both dataframes
    merged_data['Color Code'] = merged_data['Color Code'].astype(str)
    rfid_data['Color Code'] = rfid_data['Color Code'].astype(str)

    # Group the RFID data by 'DO No./Product No.', 'Color Code', 'Pack Method' and sum 'Packing Quantity'
    rfid_grouped = rfid_data.groupby(['DO No./Product No.', 'Color Code', 'Pack Method'])['Packing Quantity'].sum().reset_index()
    rfid_grouped.rename(columns={'Packing Quantity': 'RFID'}, inplace=True)

    # Merge the datasets based on specified keys
    merged_final_data = pd.merge(
        merged_data,
        rfid_grouped,
        how='left',
        left_on=['VPO No', 'Color Code', 'Pack Method'],
        right_on=['DO No./Product No.', 'Color Code', 'Pack Method']
    )

    # Drop the key columns from the right dataset
    merged_final_data.drop(columns=['DO No./Product No.'], inplace=True)

    # Ensure 'RFID' and 'CO Qty' columns are numeric, and handle NaN values
    merged_final_data['RFID'] = pd.to_numeric(merged_final_data['RFID'], errors='coerce').fillna(0)
    merged_final_data['CO Qty'] = pd.to_numeric(merged_final_data['CO Qty'], errors='coerce').fillna(0)

    # Calculate 'RFID%' and multiply by 100 to avoid percentage sign
    merged_final_data['RFID%'] = (merged_final_data['RFID'] / merged_final_data['CO Qty']).fillna(0) * 100
    merged_final_data['RFID%'] = merged_final_data['RFID%'].round(2)

    # Ensure 'Del%' is treated as a string before removing the percentage sign
    merged_final_data['Del%'] = merged_final_data['Del%'].astype(str)

    # Convert 'Del%' to numeric by removing the percentage sign and handle NaN values
    merged_final_data['Del_Dummy%'] = merged_final_data['Del%'].str.rstrip('%').astype(float).fillna(0)

    # Convert 'Min CO Sts' to numeric and handle any errors by coercing invalid values to NaN
    merged_final_data['Min CO Sts'] = pd.to_numeric(merged_final_data['Min CO Sts'], errors='coerce').fillna(0)

    # Adding the 'Status' column with the specified conditions
    def determine_status(row):
        if row['Del_Dummy%'] >= 100.0:
            return "Shipped"
        elif row['Del_Dummy%'] <= 0.0:
            return "Pending"
        elif row['Del_Dummy%'] > 0.0 and row['Del_Dummy%'] < 100.0 and row['Min CO Sts'] < 66:
            return "Short Ship"
        elif row['Del_Dummy%'] > 0.0 and row['Del_Dummy%'] < 100.0 and row['Min CO Sts'] >= 66:
            return "Short Close"    
        else:
            return ''

    merged_final_data['Status'] = merged_final_data.apply(determine_status, axis=1)

    # Replace NaN values with blanks
    merged_final_data = merged_final_data.fillna('')

    return merged_final_data

# Function to reorder columns and save the final report
def reorder_and_save_columns(final_merged_data_with_status):
    # List of columns in the desired order
    desired_order = [
        'Main Sample Code','Style No', 'Season_y', 'Year', 'Production Plan Type Name', 
        'Production Plan ID', 'VPO No', 'Item Description', 'Destination', 
        'Business Unit', 'EXF', 'Contracted Date', 'Requested Wh Date', 
        'Color Name', 'PED', 'Shipment Mode', 'MODE', 'EXF-PLN', 'ETD-PLN', 
        'POWH-PLN', 'Delays', 'Delay/Early', 'Factory – Remarks', 'CO Qty', 
        'Cum Cut Qty', 'Cut%', 'Cum Sew In Qty', 'Sewin%', 'Cum Sew In Rej Qty', 
        'Sewin Rej%', 'Cum SewOut Qty', 'Sewout%', 'Cum Sew Out Rej Qty', 
        'Sewout Rej%', 'Cum CTN Qty', 'CTN%', 'Cum CTN Rej Qty', 'RFID', 
        'RFID%', 'Delivered Qty', 'Del%', 'Excess/Short Shipped Qty', 'PCD', 
        'Status','Season_x','Min CO Sts','CO No','CPO No','Z Option','Pack Method','Schedule No',
        'MFG Schedule','Trans Reason','Req Del date','Plan Del Date ' 
    ]

    # Reorder the columns
    final_merged_data_with_status = final_merged_data_with_status[desired_order]

    # Save the modified DataFrame to a new Excel file
    output_path = r'C:\India Plants\Marketing\Shamil Report_OB\finalizedreport.xlsx'
    final_merged_data_with_status.to_excel(output_path, index=False)

    return output_path

# Function to update finalized report with delivery status data
def update_with_delivery_status(finalized_report_path, delivery_status_path):
    # Load the files
    finalized_report_df = pd.read_excel(finalized_report_path)
    delivery_status_df = pd.read_excel(delivery_status_path, engine='openpyxl')

    # Update relevant columns in the finalized report
    finalized_report_df['MODE'] = finalized_report_df.apply(
        lambda row: delivery_status_df.loc[delivery_status_df['Production Plan ID'] == row['Production Plan ID'], 'MODE'].values[0]
        if not delivery_status_df.loc[delivery_status_df['Production Plan ID'] == row['Production Plan ID'], 'MODE'].empty else row['MODE'],
        axis=1
    )

    finalized_report_df['EXF-PLN'] = finalized_report_df.apply(
        lambda row: delivery_status_df.loc[delivery_status_df['Production Plan ID'] == row['Production Plan ID'], 'EXF-PLN'].values[0]
        if not delivery_status_df.loc[delivery_status_df['Production Plan ID'] == row['Production Plan ID'], 'EXF-PLN'].empty else row['EXF-PLN'],
        axis=1
    )

    finalized_report_df['ETD-PLN'] = finalized_report_df.apply(
        lambda row: delivery_status_df.loc[delivery_status_df['Production Plan ID'] == row['Production Plan ID'], 'ETD-PLN'].values[0]
        if not delivery_status_df.loc[delivery_status_df['Production Plan ID'] == row['Production Plan ID'], 'ETD-PLN'].empty else row['ETD-PLN'],
        axis=1
    )

    finalized_report_df['POWH-PLN'] = finalized_report_df.apply(
        lambda row: delivery_status_df.loc[delivery_status_df['Production Plan ID'] == row['Production Plan ID'], 'POWH-PLN'].values[0]
        if not delivery_status_df.loc[delivery_status_df['Production Plan ID'] == row['Production Plan ID'], 'POWH-PLN'].empty else row['POWH-PLN'],
        axis=1
    )

    finalized_report_df['Factory – Remarks'] = finalized_report_df.apply(
        lambda row: delivery_status_df.loc[delivery_status_df['Production Plan ID'] == row['Production Plan ID'], 'Factory – Remarks'].values[0]
        if not delivery_status_df.loc[delivery_status_df['Production Plan ID'] == row['Production Plan ID'], 'Factory – Remarks'].empty else row['Factory – Remarks'],
        axis=1
    )

    # Remove rows where 'CO Qty' is less than zero
    finalized_report_df = finalized_report_df[finalized_report_df['CO Qty'] >= 0]

    # Update the 'Delays' field
    finalized_report_df['Delays'] = finalized_report_df.apply(
        lambda row: 0 if row['POWH-PLN'] == 0 else (pd.to_datetime(row['Requested Wh Date']) - pd.to_datetime(row['POWH-PLN'])).days,
        axis=1
    )

    # Update the 'Delay/Early' field
    finalized_report_df['Delay/Early'] = finalized_report_df.apply(
        lambda row: 'No Delay' if row['Delays'] >= 0 else 'Delay',
        axis=1
    )

    # Save the updated DataFrame to an Excel file
    output_path_updated = r'C:\India Plants\Marketing\Shamil Report_OB\finalizedreport_updated.xlsx'
    finalized_report_df.to_excel(output_path_updated, index=False)

    # Apply conditional formatting
    wb = load_workbook(output_path_updated)
    ws = wb.active

    # Apply conditional formatting for 'Delay/Early' and 'Status' columns
    apply_conditional_formatting(ws)

    # Save the workbook with formatting
    wb.save(output_path_updated)

    return output_path_updated

# Function to apply conditional formatting to the Excel sheet
def apply_conditional_formatting(ws):
    # Define the dark green and red fill colors with white text
    dark_green_fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    white_font = Font(color="FFFFFF")

    # Apply formatting to 'Delay/Early' column
    delay_early_col_idx = None
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == 'Delay/Early':
            delay_early_col_idx = col[0].column
            break

    if delay_early_col_idx is not None:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=delay_early_col_idx, max_col=delay_early_col_idx):
            for cell in row:
                if cell.value == 'No Delay':
                    cell.fill = dark_green_fill
                    cell.font = white_font
                elif cell.value == 'Delay':
                    cell.fill = red_fill
                    cell.font = white_font

    # Apply formatting to 'Status' column
    status_col_idx = None
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == 'Status':
            status_col_idx = col[0].column
            break

    if status_col_idx is not None:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=status_col_idx, max_col=status_col_idx):
            for cell in row:
                if cell.value == 'Shipped':
                    cell.fill = dark_green_fill
                    cell.font = white_font

# Streamlit app
def main():
    st.set_page_config(page_title="OB Macro", layout="wide")
    st.sidebar.title("OB Macro")
    st.sidebar.write("Upload the required files for processing.")
    
    # Upload Excel files
    uploaded_excel_1 = st.sidebar.file_uploader("Choose the first Excel file", type="xlsx")
    uploaded_csv = st.sidebar.file_uploader("Choose a CSV file", type="csv")
    uploaded_excel_2 = st.sidebar.file_uploader("Choose the second Excel file (RFID Gihan)", type="xlsx")
    uploaded_delivery_status = st.sidebar.file_uploader("Choose the Delivery Status Excel file", type="xlsx")

    st.markdown("<h2 style='text-align: center; color: #4CAF50;'>OB Macro Processing Tool</h2>", unsafe_allow_html=True)
    st.write("This tool processes multiple files, merges them, and applies updates and conditional formatting.")
    
    if uploaded_excel_1 and uploaded_csv and uploaded_excel_2 and uploaded_delivery_status:
        ob_clean_df = process_excel(uploaded_excel_1)
        spl_clean_df = process_csv(uploaded_csv)
        rfid_clean_df = process_rfid_excel(uploaded_excel_2)

        # Update the 'Production Plan ID' in the OB_clean DataFrame
        updated_df = update_production_plan_id(ob_clean_df, spl_clean_df)

        # Merge the updated OB_clean DataFrame with SPL_clean DataFrame
        merged_df = merge_dataframes(updated_df, spl_clean_df)

        # Perform final calculations and add columns
        final_df = perform_final_calculations(merged_df)

        # Add 'Color Code' to the final merged data based on 'Color Name'
        final_df_with_color_code = add_color_code(final_df)

        # Perform final merge with RFID data and add 'Status' column
        final_merged_data_with_status = final_merge_and_status(final_df_with_color_code, rfid_clean_df)

        # Reorder columns and save the final report
        finalized_report_path = reorder_and_save_columns(final_merged_data_with_status)

        # Update with delivery status data
        output_path_updated = update_with_delivery_status(finalized_report_path, uploaded_delivery_status)

        # Provide download option for the final processed data
        with open(output_path_updated, "rb") as file:
            st.download_button(
                label="Download Finalized Report with Updates",
                data=file,
                file_name="finalizedreport_updated.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

if __name__ == "__main__":
    main()
